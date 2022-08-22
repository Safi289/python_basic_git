"""
#Your first python program
print("Safi Ul Sahid")
"""
"""
#How python code get executed
print("o----")
print(" ||||")
print('*' * 10)
"""

"""
#variables

price = 10
rating = 4.9
name = 'Safi'
is_published = True
print(price)

#exercise
full_name = 'John Smith'
age = 20
is_new = True
"""

"""
#receiving input

name = input('What is your name? ')
print('Hi ' + name)

#exercise

name = input('What is your name? ')
favorite_color = input('What is your favorite color? ')
print(name + ' likes ' + favorite_color)
"""

"""
#type conversion

birth_year = input('Birth Year: ')
print(type(birth_year))
age = 2022 - int(birth_year)
print(type(age))
print(age)

#exercise

weight_lbs = input('Weight (lbs): ')
weight_kg = int(weight_lbs) * 0.45
print(weight_kg)
"""

"""
#string

course = ''' hello jon, 
 How are You?
 I am fine.
 '''
course = "Python's course for beginners"
print(course[0:])

#exercise
name = 'Jennifer'
print(name[1:-1])
"""

"""
#formatted strings
first = 'Safi'
last = 'Sahid'
message = first + ' [' + last + '] is a coder'
msg = f'{first} [{last}] is a coder'
print(msg)
"""
"""
#string methods
course = 'Python for Beginners'
print(len(course))
print(course.upper())
print(course.lower())
print(course.find('beginners'))
print(course.replace('Beginners', 'absolute beginners'))
print(course.replace('P', 'J'))
print('Python' in course)
"""

"""
# arithmetic operations
print(10 + 3)
print(10 - 3)
print(10 * 3)
print(10 / 3)
print(10 // 3)
print(10 % 3)
print(10 ** 3)

x = 10
# x = x + 3
x += 3
x -= 3
print(x)
"""
"""
# operator precedence

x = 10 + 3 * 2 ** 2
print(x)

# exercise

x = (2 + 3) * 10 - 3
print(x)
"""

"""
# math functions

x = 2.9
print(round(x))
print(abs(-2.9))

import math
print(math.ceil(2.9))
print(math.floor(2.9))
"""

"""
# if satatements

is_hot = False
is_cold = False
if is_hot:
    print("It's a hot day")
    print("Drink plenty of water.")
elif is_cold:
    print("It's a cold day")
    print("Wear warm cloths.")
else:
    print("It's a lovely day.")
print("Enjoy your day!")

# exercise

price = 1000000
has_good_credit = True

if has_good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.2 * price
print(f"Down Payment: ${down_payment}")
"""
"""
# logical operators

has_good_income = True
has_good_credit = True
has_criminal_record = False
# if has_good_income and has_good_credit:
#     print('Eligible for loan')
#
# if has_good_income or has_good_credit:
#     print('Eligible for loan')

if has_good_credit and not has_criminal_record:
    print('Eligible for loan')
"""
"""
# comparison operator

temperature = 30
if temperature > 30:
    print("It's a hot day")
else:
    print("It's not a hot day")

# exercise

name = 'John'
if len(name) < 3:
    print("Name must be at least 2 characters.")
elif len(name) > 50:
    print("Name must be maximum 50 characters.")
else:
    print("Name looks good")
"""
"""
# weight converter program

weight = int(input("Weight: "))
unit = input("(L)bs or (K)g: ")

if unit.upper() == "L":
    converted = weight * 0.45
    print(f"You are {converted} kilos")
else:
    converted = weight / 0.45
    print(f"You are {converted} pounds")
"""
"""
# while loops

i = 1
while i <= 5:
    print(i)
    print('*' * i)
    i = i + 1
print("Done")
"""
"""
# building a guessing game
secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
    guess = int(input("Guess: "))
    guess_count += 1
    if guess == secret_number:
        print("You Won!")
        break
else:
    print("Sorry, You Failed!")
"""
"""
# building the car game
command = ""
started = False

while True:
    command = input("> ").lower()
    if command == "start":
        if started:
            print("car is already started")
        else:
            started = True
            print("Car Started.....")
    elif command == "stop":
        if not started:
            print("car is already stopped")
        else:
            started = False
            print("Car Stopped.")
    elif command == "help":
        print(''' 
start - to start the car
stop - to stop the car
quit - to quit
        ''')
    elif command == "quit":
        break
    else:
        print("Sorry, I dont Understand That!")
"""
"""
# for loops
for item in 'python':
    print(item)
for item in ['safi', 'sahid', 'nur']:
    print(item)

for item in range(10):
    print(item)
for item in range(5,10):
    print(item)
for item in range(5, 10, 2):
    print(item)

# exercise
prices = [10, 20, 30]
total = 0
for price in prices:
    total += price
print(f"Total: {total}")

"""
"""
# nested loops

for x in range(4):
    for y in range(3):
        print(f"({x},{y})")

# exercise

numbers = [5, 2, 5, 2, 2]

for x_count in numbers:
    output = ''
    for count in range(x_count):
        output += 'X'
    print(output)
"""

"""
# lists
names = ['John', 'Bob', 'Mosh', 'Sarah', 'Mary']
names[0] = 'jon'
print(names[2:4])

# exercise
numbers = [2, 9, 1, 4, 3, 7]
max = numbers[0]
for number in numbers:
    if number > max:
        max = number
print(max)
"""
"""
# 2D lists
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
# matrix[0][1] = 20
# print(matrix[0][1])
for row in matrix:
    for item in row:
        print(item)
"""
"""
# list methods

numbers = [2, 1, 5, 7, 4]
# numbers.append(20)
# numbers.insert(0, 20)
# numbers.remove(5)
# numbers.clear()
# numbers.pop()
# numbers.index(1)
# numbers.sort()
# numbers.reverse()
# numbers2 = numbers.copy()
# print(numbers.index(1))
# print(numbers.count(5))
# print(numbers)

# exercise
numbers = [2, 2, 4, 6, 3, 4, 6, 1, 5, 5, 5]
uniques = []
for number in numbers:
    if number not in uniques:
        uniques.append(number)
print(uniques)
"""
"""
# tuples
numbers = (1, 2, 3)
print(numbers[0])
"""
"""
# unpacking
coordinates = (1, 2, 3)
x, y, z = coordinates
print(x)
"""

"""
# dictionaries

customer = {
    "name": "Safi Ul Sahid",
    "age": 25,
    "is_verified": True
}
# customer["name"] = "Abul"
# customer["birthdate"] = "dec 29, 93"
# print(customer["birthdate"])

# exercise
phone = input("Phone: ")
digits_mapping = {
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "Four"
}
output = ""
for ch in phone:
    output += digits_mapping.get(ch, "!") + " "
print(output)
"""
"""
# emoji converter
message = input(">")
words = message.split()
emojis = {
    ":)": "😊",
    ":(": "😒",
}
output = ""
for word in words:
    output += emojis.get(word, word) + " "
print(output)
"""
"""
# functions

def greet_user():
    print("Hi There")
    print("Good Morning Sunshine!")


print("Start")
greet_user()
print("Finish")
"""
"""
# parameters

def greet_user(first_name, last_name):
    print(f"Hi {first_name} {last_name}")
    print("Good Morning Sunshine!")

print("start")
greet_user("Safi", "Ul Sahid")
print("finish")
"""
"""
# keyword arguments

def greet_user(first_name, last_name):
    print(f"Hi {first_name} {last_name}")
    print("Good Morning Sunshine!")

print("start")
greet_user("Safi", last_name="Ul Sahid")
print("finish")
"""

"""
# return statement

def square(number):
    print(number * number)

print(square(3))
"""
"""
# creating a reusable function
def emoji_converter(message):
    words = message.split()
    emojis = {
        ":)": "😊",
        ":(": "😒",
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


message = input(">")
print(emoji_converter(message))
"""
"""
# exceptions
try:
    age = int(input("Age: "))
    income = 20000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print("Age can not be 0")
except ValueError:
    print("Invalid Value!")
"""

"""
# comments

# print ocean is blue
print("ocean is blue")

# calculates and returns a square of a number
def square(number):
    return number * number
"""
"""
# classes

class point:
    def move(self):
        print("move")
    def draw(self):
        print("draw")


point1 = point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()

point2 = point()
point2.x = 1
print(point2.x)
"""
"""
# constructors

class point:
    def __init__(self, x, y):
        self.x = x
        self.y = y
    def move(self):
        print("move")
    def draw(self):
        print("draw")

point = point(10, 20)
point.x = 11
print(point.x)

# exercise

class Person:

    def __init__(self, name):
        self.name = name
    def talk(self):
        print(f"Hi! I am {self.name}")

safi = Person("safi")
# print(safi.name)
safi.talk()

sahid = Person("sahid")
sahid.talk()
"""
"""
# Inheritance

class Mammal:
    def walk(self):
        print("walk")

class Dog(Mammal):
    def bark(self):
        print("bark")

class Cat(Mammal):
    def be_annoying(self):
        print("annoying")

dog1 = Dog()
dog1.bark()

cat1 = Cat()
cat1.be_annoying()
"""
"""
# modules
import converters
from converters import kg_to_lbs

kg_to_lbs(100)

print(converters.kg_to_lbs(70))

# exercise
import utils

numbers = [10, 3, 6, 2]
maximum = utils.find_max(numbers)
print(maximum)
"""
"""
# assignment

class LearnedTopics:
    def abc(self):
        learned_topics = ['classes', 'destructor', 'inheritance']
        return learned_topics

    def search(self, topic_name):
        matrix = self.abc()

        for row in matrix:
            for ch in row:
                for topic in topic_name:
                    if topic == ch:
                        print(row)
                    else:
                        break
                break


topic = LearnedTopics()
topicSearch = input("Search Topic Name: ")
topic.search(topicSearch)
"""


"""
class LearnedTopics:
    def abc(self):
        learned_topics = ['classes', 'destructor', 'inheritance']
        return learned_topics

    def search(self, topic_name):
        matrix = self.abc()
        count = 0
        for row in matrix:
            for ch in row:
                if ch in topic_name:
                    print(row)
                    break
                break


topic = LearnedTopics()
topicSearch = input("Search Topic Name: ")
topic.search(topicSearch)
"""

"""
import time
class LearnedTopics:

    def data_log(self, topic_name):
        seconds = time.time()
        local_time = time.ctime(seconds)
        print(f"Topic Name: '{topic_name}' is searched in {local_time}")

    def abc(self):
        learned_topics = ['classes', 'constructor', 'inheritance', 'modules']
        return learned_topics

    def search(self, topic_name):
        matrix = self.abc()
        count = 0
        for row in matrix:
            if row[0:len(topic_name)] in topic_name:
                if row[0:len(topic_name)] == topic_name:
                    print("Yes! This topic is covered.")
                    print(f"Topic Full Name: {row}")
                    break
        for row in matrix:
            if row[0:len(topic_name)] in topic_name:
                if row[0:len(topic_name)] == topic_name:
                    count = 1
        if count == 0:
            print("No match found!")


topic = LearnedTopics()
topicSearch = input("Search Topic Name: ")
topic.search(topicSearch)
topic.data_log(topicSearch)
"""
"""
# assignment

import search


topic = search.LearnedTopics()
data_log = search.Topic()

topicSearch = input("Search Topic Name: ")

topic.search(topicSearch)
data_log.data_log(topicSearch)
"""
"""
# packages

import ecommerce.shipping
from ecommerce.shipping import calc_shipping
from ecommerce import shipping

ecommerce.shipping.calc_shipping()
calc_shipping()
shipping.calc_shipping()
"""
"""
# generating random values

import random

# for i in range(3):
    # print(random.random())
    # print(random.randint(10, 20))

members = ['safi', 'sahid', 'nur']
leader = random.choice(members)
print(leader)


# exercise
import random
class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second

dice = Dice()
print(dice.roll())
"""

"""
# working with directories

from pathlib import Path

# path = Path("ecommerce")
# print(path.exists())

# path = Path("emails")
# print(path.mkdir())
# print(path.rmdir())
path = Path()
for file in path.glob('*.py'):
    print(file)
"""

# project 1 - Automation with python
"""
# import check
import sys
try:
    import openpyxl
except ImportError as iE:
    print("Python Excel module 'openpyxl' not found")
    sys.exit(1)
"""

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)
    # print(cell.value)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

process_workbook('transactions.xlsx')
